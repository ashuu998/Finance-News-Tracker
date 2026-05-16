from dotenv import load_dotenv
import os
import requests
import pandas as pd
from datetime import datetime
from textblob import TextBlob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# =====================
# API KEY LOAD
# =====================
load_dotenv()
API_KEY = os.getenv("NEWS_API_KEY")

# =====================
# STEP 1 — NEWS COLLECT
# =====================
url = "https://newsapi.org/v2/everything"

params = {
    "q"        : "stock market india",
    "language" : "en",
    "pageSize" : 10,
    "apiKey"   : API_KEY
}

response = requests.get(url, params=params)
data = response.json()

print(f"Total news received: {data['totalResults']}")

# =====================
# STEP 2 — DATA SAVE
# =====================
news_list = []

for article in data["articles"]:
    news_list.append({
        "title"      : article["title"],
        "description": article["description"],
        "source"     : article["source"]["name"],
        "published"  : article["publishedAt"],
        "url"        : article["url"]
    })

df = pd.DataFrame(news_list)
df.to_csv("news_data.csv", index=False)
print(f"Total news saved: {len(df)}")

# =====================
# STEP 3 — SENTIMENT
# =====================
def get_sentiment(text):
    if text is None:
        return "Unknown"
    score = TextBlob(text).sentiment.polarity
    if score > 0.1:
        return "Positive"
    elif score < -0.1:
        return "Negative"
    else:
        return "Neutral"

df["sentiment"] = df["title"].apply(get_sentiment)
df.to_csv("news_data.csv", index=False)

print("\n--- Sentiment Summary ---")
print(df["sentiment"].value_counts())

# =====================
# STEP 4 — EXCEL REPORT
# =====================
wb = Workbook()
ws = wb.active
ws.title = "News Report"

headers = ["Title", "Source", "Published", "Sentiment", "URL"]
ws.append(headers)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="000080")
    cell.alignment = Alignment(horizontal="center")

for _, row in df.iterrows():
    ws.append([
        row["title"],
        row["source"],
        row["published"],
        row["sentiment"],
        row["url"]
    ])

for row in ws.iter_rows(min_row=2):
    sentiment = row[3].value
    if sentiment == "Positive":
        color = "90EE90"
    elif sentiment == "Negative":
        color = "FFB6B6"
    else:
        color = "FFFFE0"
    for cell in row:
        cell.fill = PatternFill(fill_type="solid", fgColor=color)

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 40

wb.save("news_report.xlsx")
print("Excel report ready!")

# =====================
# STEP 5 — DASHBOARD
# =====================
total    = len(df)
positive = len(df[df["sentiment"] == "Positive"])
negative = len(df[df["sentiment"] == "Negative"])
neutral  = len(df[df["sentiment"] == "Neutral"])

if positive > negative:
    mood  = "BULLISH 📈"
    analysis = "Market positive!"
elif negative > positive:
    mood  = "BEARISH 📉"
    analysis = "Market negative!"
else:
    mood  = "NEUTRAL ➡️"
    analysis = "Market stable!"

print("\n" + "="*45)
print("   FINANCIAL NEWS DASHBOARD")
print("="*45)